import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import load_workbook
from form import create_form_frame
from data_page import create_data_frame

# Variables globales para almacenar empresas y alumnos
empresas = []
num_empresas = 0
alumnos = {}

hora_general = "15:00"
tiempo_entrevista_general = "5"
tiempo_cambio_general = "1"

def add_alumno_to_schedule(empresas, alumnos, nombre_alumno, nombre_empresa):
    max_positions = 50  # Número máximo de filas en la tabla
    
    for position in range(max_positions):
        conflict = False
        
        # Verificar si la posición actual está ocupada por el mismo alumno en alguna empresa
        for empresa in empresas:
            if len(alumnos[empresa]) > position:
                if alumnos[empresa][position] == nombre_alumno:
                    conflict = True
                    break

        # Si no hay conflicto, añadimos al alumno a la empresa especificada
        if not conflict:
            while len(alumnos[nombre_empresa]) <= position:
                alumnos[nombre_empresa].append("null")
            if alumnos[nombre_empresa][position] == "null":
                alumnos[nombre_empresa][position] = nombre_alumno
                return  # Alumno añadido exitosamente

    return "No hay posiciones libres disponibles para añadir al alumno."

def import_from_xlsx():
    global empresas, alumnos, num_empresas

    file_path = filedialog.askopenfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        wb = load_workbook(file_path)
        ws = wb.active

        # Obtener la lista de empresas
        empresas = [cell.value for cell in ws[1] if cell.value != "Horas"]
        num_empresas = len(empresas)
        alumnos = {empresa: [] for empresa in empresas}

        # Leer los datos de las filas
        for row in ws.iter_rows(min_row=2, values_only=True):
            hora = row[0]  # Hora es la primera columna
            for col_index, value in enumerate(row[1:], start=0):
                if value and value != "null":
                    empresa = empresas[col_index]
                    add_alumno_to_schedule(empresas, alumnos, value.lower(), empresa)

        show_data_page(hora_general, tiempo_entrevista_general, tiempo_cambio_general)


# Función para cambiar a la pantalla del formulario
def show_form():
    main_frame.pack_forget()
    data_frame.pack_forget()
    form_frame.pack(fill='both', expand=True)

# Función para volver al menú principal
def back_to_main():
    form_frame.pack_forget()
    data_frame.pack_forget()
    main_frame.pack(fill='both', expand=True)

# Función para mostrar la página de datos
def show_data_page(hora='', tiempo_entrevista='', tiempo_cambio=''):
    global hora_general
    global tiempo_entrevista_general
    global tiempo_cambio_general

    if hora != '':
        hora_general = hora
    if tiempo_entrevista != '':
        tiempo_entrevista_general = tiempo_entrevista
    if tiempo_cambio != '':
        tiempo_cambio_general = tiempo_cambio

    form_frame.pack_forget()
    main_frame.pack_forget()
    for widget in data_frame.winfo_children():
        widget.destroy()
    create_data_frame(data_frame, back_to_main, hora_general, tiempo_entrevista_general, tiempo_cambio_general, empresas, num_empresas, alumnos, add_empresa, add_alumno)
    data_frame.pack(fill='both', expand=True)

# Función para añadir una empresa
def add_empresa():
    global num_empresas
    def save_empresa():
        global num_empresas
        nombre_empresa = empresa_entry.get()
        if nombre_empresa.lower() in empresas:
            messagebox.showerror("Error", "Empresa ya agregada.")
        elif nombre_empresa and num_empresas < 50:  # Limitar a 50 empresas
            empresas.append(nombre_empresa.lower())
            alumnos[nombre_empresa] = []
            num_empresas += 1
            empresa_window.destroy()
            show_data_page()

    empresa_window = tk.Toplevel(root)
    empresa_window.title("Añadir Empresa")
    ttk.Label(empresa_window, text="Nombre de la Empresa:").pack(pady=10)
    empresa_entry = ttk.Entry(empresa_window)
    empresa_entry.pack(pady=5)
    ttk.Button(empresa_window, text="Guardar", command=save_empresa).pack(pady=10)

# Función para añadir un alumno
def add_alumno():
    def save_alumno():
        nombre_alumno = alumno_entry.get()
        nombre_empresa = empresa_combobox.get()
        if nombre_alumno.lower() in alumnos[nombre_empresa]:
            messagebox.showerror("Error", "Alumno ya agregado.")
        elif nombre_alumno and nombre_empresa in empresas:
            # a la hora de agregarlo tengo que controlar que no se choque con otros. Puedo agregar uno que sea null.
            add_alumno_to_schedule(empresas, alumnos, nombre_alumno.lower(), nombre_empresa)
            alumno_window.destroy()
            show_data_page()

    alumno_window = tk.Toplevel(root)
    alumno_window.title("Añadir Alumno")
    ttk.Label(alumno_window, text="Nombre del Alumno:").pack(pady=10)
    alumno_entry = ttk.Entry(alumno_window)
    alumno_entry.pack(pady=5)
    ttk.Label(alumno_window, text="Nombre de la Empresa:").pack(pady=10)
    empresa_combobox = ttk.Combobox(alumno_window, values=empresas[:num_empresas])
    empresa_combobox.pack(pady=5)
    ttk.Button(alumno_window, text="Guardar", command=save_alumno).pack(pady=10)

# Crear la ventana principal
root = tk.Tk()
root.title("1to1 APP")
root.geometry("800x600")

# Crear el menú principal
menubar = tk.Menu(root)
root.config(menu=menubar)

# Crear un menú de opciones
file_menu = tk.Menu(menubar, tearoff=0)
file_menu.add_command(label="Nuevo Horario", command=show_form)
file_menu.add_command(label="Importar XLSX", command=import_from_xlsx)
file_menu.add_command(label="Salir", command=root.quit)
menubar.add_cascade(label="Archivo", menu=file_menu)

# Crear otro menú de opciones (desplegable)
edit_menu = tk.Menu(menubar, tearoff=0)
edit_menu.add_command(label="Añadir Empresa", command=add_empresa)
edit_menu.add_command(label="Añadir Alumno", command=add_alumno)
menubar.add_cascade(label="Editar", menu=edit_menu)

# Frame del menú principal
main_frame = ttk.Frame(root)
main_frame.pack(fill='both', expand=True)

# Crear el frame del formulario llamando a una función del archivo form.py
form_frame = create_form_frame(root, back_to_main, show_data_page)

# Crear el frame de la página de datos
data_frame = ttk.Frame(root)

# Contenido del menú principal
main_label = ttk.Label(main_frame, text="Menú Principal", font=("Helvetica", 16))
main_label.pack(pady=20)

nuevo_horario_button = ttk.Button(main_frame, text="Nuevo Horario", command=show_form)
nuevo_horario_button.pack(pady=10)

import_xlsx_button = ttk.Button(main_frame, text="Importar XLSX", command=import_from_xlsx)
import_xlsx_button.pack(pady=10)

dummy_button = ttk.Button(main_frame, text="Botón sin acción")
dummy_button.pack(pady=10)

# Iniciar el bucle principal de la interfaz
root.mainloop()
